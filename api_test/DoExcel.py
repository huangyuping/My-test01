import os
from openpyxl import load_workbook
import xlrd

class DoExcel:

    def __init__(self,excelpath):
        self.wb = load_workbook(excelpath)
        self.excelpath=excelpath
        self.sheet_apitest = self.wb['api_test']
        self.sheet_phone = self.wb['api_phone']

    #读取phone表格内容
    def Excel_phone(self):
        init_datas={}
        for index in  range(2,self.sheet_phone.max_row+1):
            key=self.sheet_phone.cell(row=index,column=1).value

            init_datas[key]=self.sheet_phone.cell(row=index,column=2).value
           # init_datas['${phone_b}']=str(int(init_datas['${phone_b}'])+1)

            return init_datas
    #修改phone表格手机号
    def modify_phone(self):
        res=int(self.sheet_phone.cell(row=2,column=2).value)
        #print(res)
        self.sheet_phone.cell(row=2,column=2).value=str(res+1)
        #print(self.sheet_phone.cell(row=2,column=2).value)
    #保存phone表格
    def save_excelFile(self,):
        self.wb.save(self.excelpath)


    #读取api_test表格
    def Excel_api_test(self):
        all_case_datas=[]
        for i in  range(2,self.sheet_apitest.max_row+1):
            case_data={}

            case_data['id']=self.sheet_apitest.cell(row=i,column=1).value
            case_data['api_name']=self.sheet_apitest.cell(row=i,column=2).value
            case_data["url"]=self.sheet_apitest.cell(row=i,column=3).value
            request_data=self.sheet_apitest.cell(row=i,column=4).value
            init_datas=self.Excel_phone()
            #替换request_data 中的手机号
            if request_data is not None and len(init_datas)>0:
                for key , value in init_datas.items():
                    if request_data.find(key)!=-1:
                        request_data = request_data.replace(str(key),str(value))

            case_data["request_data"]=request_data
            case_data["method"]=self.sheet_apitest.cell(row=i,column=5).value
            case_data["expected_data"]=self.sheet_apitest.cell(row=i,column=6).value
            if self.sheet_apitest.cell(row=i,column=5).value is not None:
                case_data['related_exp']=self.sheet_apitest.cell(row=i,column=7).value
            all_case_datas.append(case_data)
        return all_case_datas





# a=DoExcel('/Users/xiaolongxia/PycharmProjects/api_test/text.xlsx')
# print(a.Excel_api_test())
# print(a.Excel_phone())
# a.modify_phone()
# a.save_excelFile()


#print(DoExcel.Excel('/Users/xiaolongxia/PycharmProjects/api_test/text.xlsx'))
